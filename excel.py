import os
import sys
import pyexcel as p
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill

folder_path = os.getcwd()
xl_sx_files = []
xl_s_files = []
summary_files = []
need_compare_files = []
compare_summary_sheet = "核对结果.xlsx"

# 存储需要的列数
Compare_indices = {'姓名': None, '全勤': None, '出勤天数': None, '平时': None, '周末': None, '法定': None, '晚餐补贴': None, '迟到': None, '事假（天）': None, '病假（天）': None, '年假（天）': None}
# 存储系统汇总表需要的列数
system_indices = {'姓名': None, '全勤': None, '实出勤天数': None, '加班1.5': None, '加班2.0': None, '加班3.0': None, '夜班次数': None, '迟到次数': None, '事假天数': None, '病假天数': None, '年休假天数': None}
# 存储汇总表右边的标题
titleList_right = ['姓名', '全勤', '实出勤天数', '平时', '周末', '法定', '晚餐补贴', '迟到', '事假', '病假', '年假']
# 存储汇总表左边的标题
titleList_left = ['姓名', '全勤', '实出勤天数', '加班1.5', '加班2.0', '加班3.0', '夜班次数', '迟到次数', '事假天数', '病假天数', '年休假天数']
# 存储要复制的信息
data = []
data_name = []


# 获取所需要的行数（根据名字）
def from_name_get_need_row(indices, file_name):
    source_wb = load_workbook(os.path.join(folder_path, file_name), data_only=True)
    source_sheet = source_wb.active

    for row in source_sheet.iter_rows():
        for cell in row:
            for header in indices.keys():
                if header in str(cell.value):
                    indices[header] = cell.row

    source_wb.close()
    # print(indices)
    return indices


# 获取所需要的行数（根据工号列，找到包含2200的行）
def get_need_row(file_name):
    job_cell = 0
    source_wb = load_workbook(os.path.join(folder_path, file_name), data_only=True)
    source_sheet = source_wb.active

    job_row = []
    for row in source_sheet.iter_rows():
        for cell in row:
            if '工号' in str(cell.value):
                job_cell = row.index(cell)
    for row in source_sheet.iter_rows():
        for cell in row:
            if '2200' in str(cell.value) and job_cell == row.index(cell):
                job_row.append(cell.row)

    source_wb.close()
    # print(job_row)
    return job_row


# 根据获取需要的列数（根据所需要的标题，获取列数，只遍历rows_with_job最后一行之前内容）
def get_need_cell(rows_with_job, indices, file_name):
    source_wb = load_workbook(os.path.join(folder_path, file_name), data_only=True)
    source_sheet = source_wb.active

    for row in source_sheet.iter_rows(max_row=max(rows_with_job), values_only=True):
        for cell in row:
            for header in indices.keys():
                if header in str(cell):
                    indices[header] = row.index(cell)+1
    indices = {header: 1000 if value is None else value for header, value in indices.items()}
    source_wb.close()
    # print(indices)
    pass


def compare_fun(workbook, sheet_name, file_name):
    # 获取所需要的行数
    rows_with_job = get_need_row(file_name)

    # 获取需要的列数
    if '考勤汇总' in str(file_name):
        indices = system_indices
        data_len = 1
    else:
        indices = Compare_indices
        title_list = titleList_right
        data_len = len(title_list)
    get_need_cell(rows_with_job, indices, file_name)

    # 给表创建一个标题
    ws = workbook[sheet_name]
    for i, value in enumerate(titleList_left):
        ws.cell(row=1, column=i + 1, value=value)
    for i, value in enumerate(titleList_right):
        ws.cell(row=1, column=i + 1 + len(titleList_left) + 1, value=value)

    # 获取单元格内容并赋值到另一个表
    source_wb = load_workbook(os.path.join(folder_path, file_name), data_only=True)
    source_sheet = source_wb.worksheets[0]

    start_row = 2
    # 循环把一行单元格数据，放到data，然后粘贴到另一个表
    for row in rows_with_job:
        for cell in indices.values():
            if cell == list(indices.values())[0]:
                data_name.append(source_sheet.cell(row=row, column=cell).value)
            data.append(source_sheet.cell(row=row, column=cell).value)
        # print(data)
        for i, value in enumerate(data):
            if value is None:
                ws.cell(row=start_row, column=i + 1 + data_len + 1, value=0)
            else:
                ws.cell(row=start_row, column=i + 1 + data_len + 1, value=value)
        start_row += 1
        data.clear()
    source_wb.close()
    pass
    data_name_indices = {item: None for item in data_name}
    # print(data_name_indices)
    # 根据名字获取系统表的行列
    rows_with_name_summary = from_name_get_need_row(data_name_indices, summary_files[0]).values()
    get_need_cell(rows_with_name_summary, system_indices, summary_files[0])

    # 打开总表
    source_wb = load_workbook(os.path.join(folder_path, summary_files[0]), data_only=True)
    source_sheet = source_wb.worksheets[0]

    start_row = 2
    # 循环把一行单元格数据，放到data，然后粘贴到另一个表
    for row in rows_with_name_summary:
        for cell in system_indices.values():
            data.append(source_sheet.cell(row=row, column=cell).value)
        # print(data)
        for i, value in enumerate(data):
            ws.cell(row=start_row, column=i + 1, value=value)
        start_row += 1
        data.clear()
    source_wb.close()
    pass
    data_name.clear()
    workbook.save(os.path.join(folder_path, compare_summary_sheet))


# 去掉路径和文件名后缀，只保留文件名
def get_filename_without_extension(file_path):
    # 去掉路径
    file_name = os.path.basename(file_path)
    # 去掉后缀
    return os.path.splitext(file_name)[0]


# 获取需要对比的文件名，和总文件名
def get_xls_or_sx_summary_files():
    global need_compare_files
    for file in os.listdir(folder_path):
        if file.endswith(".xls"):
            file = get_filename_without_extension(file)
            # print(file)
            p.save_book_as(file_name=(os.path.join(folder_path, (file + '.xls'))), dest_file_name=(os.path.join(folder_path, (file + '.xlsx'))))

    for file in os.listdir(folder_path):
        if file == compare_summary_sheet:
            continue
        if file.endswith(".xlsx"):
            if "考勤汇总" in file:
                summary_files.append(file)
            else:
                if file.endswith(".xlsx"):
                    xl_sx_files.append(file)
                else:
                    xl_s_files.append(file)
    need_compare_files = xl_sx_files + xl_s_files
    print("需要核对的表：", need_compare_files)
    print("系统导出的总表：", summary_files)


# 创建工作簿
def compare_summary_file_create():
    try:
        work = load_workbook(os.path.join(folder_path, compare_summary_sheet))
        print("核对结果.xlsx已存在！")
        ret = input("输入0或1：   0->退出程序   1->覆盖")
        if ret == '1':
            pass
        elif ret == '0':
            work.close()
            sys.exit()
    except FileNotFoundError:
        work = Workbook()
        print("创建工作簿：", compare_summary_sheet)
    return work


# 创建核对工作表
def compare_summary_sheet_create(workbook, file_name):
    # 获取所有工作表名字
    sheet_names = workbook.sheetnames
    # print("工作簿里的表有：", sheet_names)
    # 获取要创建的工作表名字（去掉文件名后缀）
    sheet_name = get_filename_without_extension(file_name) + "核对"
    # 不存在则创建工作表
    if sheet_name in sheet_names:
        print("工作表已存在：", sheet_name)
        # workbook.close()
        # return 0
    else:
        workbook.create_sheet(sheet_name)
        sheet_names = workbook.sheetnames
        print("创建工作表", sheet_name)

    # 删除空表
    for delSheet in sheet_names:
        if 'Sheet' in delSheet:
            wb.remove(wb[delSheet])
    print("目前存在表：", sheet_names)
    print("正在复制数据中。。。")
    compare_fun(workbook, sheet_name, file_name)
    return 1


def compare_summary_fun():
    last_wb = load_workbook(os.path.join(folder_path, compare_summary_sheet))

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

    # print("into compare_summary_fun")
    for sheet in last_wb.worksheets:
        for row in sheet.iter_rows(min_row=2):
            # 对比每一对列
            for i in range(11):  # 从A列到K列，总共11列
                if row[i].value != row[i + 12].value:  # A列和M列的索引差为12，B列和N列的索引差为12，以此类推
                    # 如果不一致，将整行标黄
                    # print(row[i].value, row[i + 12].value)
                    for cell in row:
                        cell.fill = yellow_fill
                    # 将不一致的单元格标红
                    row[i].fill = red_fill
                    row[i + 12].fill = red_fill
    last_wb.save(os.path.join(folder_path, compare_summary_sheet))


get_xls_or_sx_summary_files()
wb = compare_summary_file_create()
for need_compare_file in need_compare_files:
    compare_summary_sheet_create(wb, need_compare_file)
compare_summary_fun()
input("对比完成，按任意键结束。。")